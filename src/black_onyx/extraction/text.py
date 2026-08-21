"""Text extraction from files — dispatch by extension with encoding detection."""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import zipfile
from email import policy
from email.parser import BytesParser
from xml.etree import ElementTree

import chardet
from bs4 import BeautifulSoup

from black_onyx.models.enums import (
    HTML_EXTENSIONS,
    PDF_EXTENSIONS,
    TEXT_EXTENSIONS,
    TEXTRACT_EXTENSIONS,
    detect_file_type,
    FileType,
)

logger = logging.getLogger(__name__)


def _detect_encoding(filepath: str) -> str:
    """Detect file encoding using chardet.

    Args:
        filepath: Path to the file.

    Returns:
        Detected encoding string (defaults to 'utf-8' if detection fails).
    """
    try:
        with open(filepath, "rb") as f:
            raw_data = f.read(1024)
        result = chardet.detect(raw_data)
        encoding = result.get("encoding") or "utf-8"
        logger.debug(f"Detected encoding: {encoding} for file: {filepath}")
        return encoding
    except Exception as e:
        logger.warning(f"Encoding detection failed for {filepath}: {e}, defaulting to utf-8")
        return "utf-8"


def _extract_html(content: str) -> str:
    """Extract text from HTML content using BeautifulSoup."""
    soup = BeautifulSoup(content, "html.parser")
    return soup.get_text(separator=" ", strip=True)


def _extract_pdf(filepath: str) -> str:
    """Extract text from a PDF file using pdfminer."""
    from pdfminer.high_level import extract_text
    return extract_text(filepath)


def _extract_csv(content: str) -> str:
    """Extract text from CSV content."""
    reader = csv.reader(io.StringIO(content))
    return " ".join(" ".join(row) for row in reader)


def _extract_tsv(content: str) -> str:
    """Extract text from TSV content."""
    reader = csv.reader(io.StringIO(content), delimiter="\t")
    return " ".join(" ".join(row) for row in reader)


def _extract_json(content: str) -> str:
    """Extract text from JSON content."""
    try:
        data = json.loads(content)
        return json.dumps(data, ensure_ascii=False)
    except json.JSONDecodeError as e:
        logger.warning(f"JSON parse error: {e}")
        return content


def _extract_jsonl(content: str) -> str:
    """Extract text from JSONL (JSON Lines) content."""
    parts: list[str] = []
    for line in content.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
            parts.append(json.dumps(obj, ensure_ascii=False))
        except json.JSONDecodeError:
            parts.append(line)
    return " ".join(parts)


def _extract_xml(content: str) -> str:
    """Extract text from XML content using BeautifulSoup."""
    soup = BeautifulSoup(content, "xml")
    return soup.get_text(separator=" ", strip=True)


def _extract_xlsx(filepath: str) -> str:
    """Extract displayed cell values from an XLSX workbook."""
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        values = [str(cell) for sheet in workbook.worksheets for row in sheet.iter_rows(values_only=True)
                  for cell in row if cell is not None]
        workbook.close()
        return " ".join(values)
    except Exception as exc:
        logger.debug("openpyxl extraction failed, using OpenXML fallback: %s", exc)
        return _extract_openxml(filepath)


def _extract_openxml(filepath: str) -> str:
    """Extract text nodes from DOCX, PPTX, XLSX, ODT, or ODS ZIP/XML containers."""
    parts: list[str] = []
    with zipfile.ZipFile(filepath) as archive:
        for name in sorted(archive.namelist()):
            if not name.endswith(".xml") or not name.startswith(("word/", "ppt/", "xl/", "content.xml")):
                continue
            root = ElementTree.fromstring(archive.read(name))
            parts.extend(value.strip() for value in root.itertext() if value.strip())
    return " ".join(parts)


def _extract_eml(filepath: str) -> str:
    with open(filepath, "rb") as source:
        message = BytesParser(policy=policy.default).parse(source)
    parts = [str(message.get(name, "")) for name in ("From", "To", "Cc", "Subject")]
    for item in message.walk():
        if item.get_content_type() in {"text/plain", "text/html"}:
            content = item.get_content()
            parts.append(_extract_html(content) if item.get_content_type() == "text/html" else content)
    return " ".join(parts)


def _extract_textract(filepath: str) -> str | None:
    """Extract text using textract as a fallback for unsupported formats.

    Args:
        filepath: Path to the file.

    Returns:
        Extracted text string, or None if textract is not installed or fails.
    """
    try:
        import textract
        text = textract.process(filepath).decode("utf-8", errors="replace")
        return text
    except ImportError:
        logger.debug("textract not installed; cannot extract this file type")
        return None
    except Exception as e:
        logger.warning(f"textract extraction failed for {filepath}: {e}")
        return None


def extract_text_from_file(filepath: str) -> str | None:
    """Extract text from a file, dispatching by file extension.

    Supports: HTML, PDF, TXT, CSV, TSV, JSON, JSONL, XML, and other formats
    via textract (if installed as optional dependency).

    Args:
        filepath: Path to the file to extract text from.

    Returns:
        Extracted text as a string, or None if extraction fails.
    """
    if not os.path.exists(filepath):
        logger.error(f"File not found: {filepath}")
        return None

    ext = os.path.splitext(filepath)[1].lower()
    file_type = detect_file_type(filepath)

    logger.debug(f"Processing file: {filepath} (extension: {ext}, type: {file_type})")

    try:
        # Binary/container formats are handled before text decoding.
        if file_type == FileType.PDF or ext in PDF_EXTENSIONS:
            return _extract_pdf(filepath)
        if ext == ".xlsx":
            return _extract_xlsx(filepath)
        if ext in {".docx", ".pptx", ".odt", ".ods"}:
            return _extract_openxml(filepath)
        if ext == ".eml":
            return _extract_eml(filepath)

        # For text-based formats, read the file with detected encoding
        encoding = _detect_encoding(filepath)
        with open(filepath, "r", encoding=encoding, errors="replace") as f:
            content = f.read()

        if file_type == FileType.HTML or ext in HTML_EXTENSIONS:
            return _extract_html(content)

        if ext == ".csv":
            return _extract_csv(content)
        if ext == ".tsv":
            return _extract_tsv(content)
        if ext == ".json":
            return _extract_json(content)
        if ext in {".jsonl", ".ndjson"}:
            return _extract_jsonl(content)
        if ext == ".xml":
            return _extract_xml(content)

        # Plain text formats
        if ext in TEXT_EXTENSIONS:
            return content

        # Try textract for other formats (.doc, .docx, .pptx, .xlsx, etc.)
        if ext in TEXTRACT_EXTENSIONS:
            return _extract_textract(filepath)

        # Unknown extension — try textract as last resort
        text = _extract_textract(filepath)
        if text is not None:
            return text

        logger.warning(f"Unsupported file extension: {ext} for {filepath}")
        return None

    except Exception as e:
        logger.error(f"Error extracting text from {filepath}: {e}")
        return None
